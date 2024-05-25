import sqlite3 as sq
from create_bot import dp, bot
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


async def sql_start():
    global base, cur
    base = sq.connect('pass.db')
    cur = base.cursor()
    if base:
        print('Data base connected')
    base.execute('CREATE TABLE IF NOT EXISTS passes(id INTEGER PRIMARY KEY AUTOINCREMENT, company, phone, '
                 'user_id INTEGER, date DATE, car_type, brand, car_number, status)')
    base.execute('CREATE TABLE IF NOT EXISTS rent(company TEXT)')
    base.execute('CREATE TABLE IF NOT EXISTS users(company TEXT, phone INTEGER, '
                 'user_id INTEGER)')
    # base.execute('CREATE UNIQUE INDEX IF NOT EXISTS name_product ON basket (user_id, name)')
    # base.execute('CREATE TABLE IF NOT EXISTS users(user_id PRIMARY KEY, firstname, username,'
    #              ' phone_number, delivery_address, pay)')
    # base.execute('CREATE TABLE IF NOT EXISTS ids(user_id PRIMARY KEY)')
    base.commit()
    print('таблицы добавлены')


async def sql_add_company(company):
    cur.execute("INSERT OR IGNORE INTO rent VALUES (?)", (company,))
    base.commit()


async def sql_add_pass(data):
    cur.execute("INSERT INTO passes(company, phone, user_id, date, car_type, brand, car_number, status) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)", tuple(data.values()))
    base.commit()


async def sql_list_company(message):
    for i in cur.execute("SELECT * FROM rent").fetchall():
        await bot.send_message(message.from_user.id, i[0], reply_markup=InlineKeyboardMarkup(row_width=1).
                               add(InlineKeyboardButton('Удалить', callback_data=f'del {i[0]}')))


async def sql_list_company2():
    cur.execute("SELECT * FROM rent")
    rows = cur.fetchall()
    return rows




async def sql_add_person(company, phone, user_id):
    cur.execute("INSERT OR IGNORE INTO users VALUES (?, ?, ?)", (company, phone, user_id))
    base.commit()


async def sql_del_company(data):
    # cur.execute("DELETE FROM passes WHERE company == ?", (data,))
    cur.execute("DELETE FROM rent WHERE company == ?", (data,))
    cur.execute("DELETE FROM users WHERE company == ?", (data,))
    base.commit()


async def sql_del_person(data):
    cur.execute("DELETE FROM users WHERE phone == ?", (data,))
    base.commit()


async def sql_list_persons(message):
    for i in cur.execute("SELECT * FROM rent").fetchall():
        await bot.send_message(message.from_user.id, f"{i[0]}\nИмя: {i[1]}\nТелефон: {i[1]}",
                               reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton('Удалить',
                                                                                            callback_data=f'delt {i}')))


# поиск номер телефона
async def search_phone_number(phone):
    cur.execute("SELECT * FROM users WHERE phone=?", (phone,))
    row = cur.fetchone()
    return row


async def add_user_id(user_id, phone):
    cur.execute("UPDATE users SET user_id = ? WHERE phone = ?", (user_id, phone))
    base.commit()


async def search_user_id(user_id):
    cur.execute("SELECT * FROM users WHERE user_id=?", (user_id,))
    row = cur.fetchone()
    print(row)
    return row


async def sql_update_status(status, user_id):
    cur.execute("SELECT id FROM passes WHERE user_id = ? AND status = ? ORDER BY id DESC LIMIT 1",(user_id, 'Анкета отправлена'))
    row = cur.fetchone()
    if row:
        last_id = row[0]
        cur.execute("UPDATE passes SET status = ? WHERE id = ?", (status, last_id))
        base.commit()
    # cur.execute("UPDATE passes SET status = ? WHERE user_id = ? AND status = ? ORDER BY id DESC LIMIT 1",
    #             (status, user_id, 'Анкета отправлена'))
    # base.commit()


async def sql_excel(chat_id, status, start_date, end_date):
    normal_style = NamedStyle(name="normal")
    normal_style.alignment = Alignment(horizontal='center')
    normal_style.font = Font(bold=False)

    # Подключение к базе данных SQLite
    # Загрузка данных из базы данных в DataFrame
    query = ("SELECT car_type, brand, car_number, date, phone, company FROM passes WHERE status = ? AND date BETWEEN ? "
             "AND ? ORDER BY company, date")

    # Передаем параметры запроса (статус и диапазон дат) при выполнении запроса
    df = pd.read_sql_query(query, base, params=(status, start_date, end_date))

    # Форматирование даты в столбце "date"
    df['date'] = pd.to_datetime(df['date']).dt.strftime('%d.%m.%Y')

    # Создание нового Excel файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Passes"

    # Установка заголовков столбцов
    headers = ['№', 'Вид ТС', 'Марка', 'Гос. номер', 'Дата', 'Заказчик']

    # Запись данных в Excel файл с группировкой по компаниям и добавлением разделителей
    current_row = 1  # Изменено начальное значение на 1
    for company, group in df.groupby('company'):
        # Записываем название компании
        company_row_start = current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        company_cell = ws.cell(row=current_row, column=1)
        company_cell.value = company
        company_cell.font = Font(bold=True, size=20)  # Увеличиваем размер шрифта
        company_cell.alignment = Alignment(horizontal='center')
        company_cell.fill = PatternFill(start_color="acb78e", end_color="acb78e", fill_type="solid")  # Зеленый фон
        # Увеличиваем высоту строки в 3 раза
        ws.row_dimensions[current_row].height = company_cell.font.size * 1.5
        current_row += 1

        # Записываем заголовки сразу после названия компании
        for col, header in enumerate(headers, start=1):
            header_cell = ws.cell(row=current_row, column=col, value=header)
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)

        current_row += 1

        # Записываем данные для каждой компании
        for index, row in group.iterrows():
            ws.append([current_row - company_row_start - 1, *list(row)[:-1]])  # Изменено начальное значение на 0
            current_row += 1

    # Применяем выравнивание к каждой ячейке
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=current_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Добавляем границы для всех строк
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=current_row, max_col=len(headers)):
        for cell in row:
            cell.border = border

    # Форматирование столбцов по длине текста
    for col, _ in enumerate(headers, start=1):
        max_length = 0
        for cell in ws[f"{get_column_letter(col)}"]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Сохранение изменений и закрытие файла Excel
    wb.save('отчет.xlsx')

    # Отправка файла пользователю
    with open('отчет.xlsx', 'rb') as file:
        await bot.send_document(chat_id, file)